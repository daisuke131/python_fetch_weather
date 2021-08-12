import openpyxl

from common.beautifulsoup import fetch_soup

WEATHER_URL = "https://weather.goo.ne.jp/past/770/{year}{month}00/"


class FetchWeather:
    def __init__(self, excel_path: str) -> None:
        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(self.excel_path)
        self.ws = self.wb.worksheets[1]

    def fetch_weather_data(self):
        month_day_weather_row = 6
        before_month = 0
        table_row = 1
        year = int(self.ws.cell(row=1, column=13).value)

        while True:
            try:
                month = int(self.ws.cell(row=month_day_weather_row, column=1).value)
                if not month:
                    break
                # 前が12月で今が１月なら年+1
                if before_month == "12" and month == "1":
                    year += 1
                day = int(self.ws.cell(row=month_day_weather_row, column=2).value)
                wather_cell = self.ws.cell(row=month_day_weather_row, column=4)
                temperature_higher_cell = self.ws.cell(
                    row=month_day_weather_row, column=5
                )
                temperature_lower_cell = self.ws.cell(
                    row=month_day_weather_row, column=6
                )

                if not before_month == month:
                    soup = fetch_soup(
                        WEATHER_URL.format(year=year, month=f"{month:02}")
                    )
                    (
                        temperature_higher_list,
                        temperature_lower_list,
                        weather_list,
                    ) = self.weather_list(soup)

                wather_cell.value = weather_list[day - 1]
                temperature_higher_cell.value = int(temperature_higher_list[day - 1])
                temperature_lower_cell.value = int(temperature_lower_list[day - 1])

                before_month = month

                if table_row == 7:
                    table_row = 1
                    month_day_weather_row += 13
                else:
                    table_row += 1
                    month_day_weather_row += 6
            except Exception:
                break

        self.wb.save(self.excel_path)

    def weather_list(self, soup):
        table = soup.findAll("table", {"class": "t01 past01 mb10"})[0]
        rows = table.findAll("tr")
        day_row = 1
        temperature_row = 2
        weather_row = 3
        temperature_higher_list = []
        temperature_lower_list = []
        weather_list = []
        while True:
            try:
                for td in rows[temperature_row].select("td"):
                    if not td.select_one(".red").get_text() == "-":
                        temperature_higher_list.append(
                            td.select_one(".red").get_text()[:2]
                        )
                        temperature_lower_list.append(
                            td.select_one(".blue").get_text()[:2]
                        )
                for td in rows[weather_row].select("td"):
                    if not td.get_text() == "-":
                        weather_list.append(td.select_one("img").get("alt")[0])
                day_row += 6
                temperature_row += 6
                weather_row += 6
            except Exception:
                break
        return temperature_higher_list, temperature_lower_list, weather_list


def fetch_weather(excel_path: str):
    fc = FetchWeather(excel_path)
    fc.fetch_weather_data()
